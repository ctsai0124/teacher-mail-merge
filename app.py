import os
import re
import zipfile
import subprocess
import tempfile
import shutil
import xml.etree.ElementTree as etree

etree.register_namespace('w', 'http://schemas.openxmlformats.org/wordprocessingml/2006/main')
etree.register_namespace('w14', 'http://schemas.microsoft.com/office/word/2010/wordml')
etree.register_namespace('w15', 'http://schemas.microsoft.com/office/word/2012/wordml')
etree.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
etree.register_namespace('wp', 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing')
etree.register_namespace('mc', 'http://schemas.openxmlformats.org/markup-compatibility/2006')
etree.register_namespace('wpc', 'http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas')
etree.register_namespace('cx', 'http://schemas.microsoft.com/office/drawing/2014/chartex')
etree.register_namespace('wne', 'http://schemas.microsoft.com/office/word/2006/wordml')
etree.register_namespace('wps', 'http://schemas.microsoft.com/office/word/2010/wordprocessingShape')
etree.register_namespace('wpg', 'http://schemas.microsoft.com/office/word/2010/wordprocessingGroup')
etree.register_namespace('wpi', 'http://schemas.microsoft.com/office/word/2010/wordprocessingInk')
etree.register_namespace('xml', 'http://www.w3.org/XML/1998/namespace')
from flask import Flask, request, send_file, jsonify, send_from_directory
from pypdf import PdfReader, PdfWriter
import openpyxl

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024

W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'


def replace_mergefields(xml_bytes, data):
    tree = etree.fromstring(xml_bytes)
    for p in tree.findall('.//{%s}p' % W):
        runs = list(p)
        i = 0
        while i < len(runs):
            el = runs[i]
            fb = el.find('.//{%s}fldChar' % W)
            if fb is not None and fb.get('{%s}fldCharType' % W) == 'begin':
                fname = None; j = i+1; druns = []; ind = False; eidx = j
                while j < len(runs):
                    r = runs[j]
                    fc2 = r.find('.//{%s}fldChar' % W)
                    it  = r.find('.//{%s}instrText' % W)
                    if it is not None:
                        m = re.search(r'MERGEFIELD\s+(\S+)', it.text or '')
                        if m: fname = m.group(1)
                    if fc2 is not None:
                        ft = fc2.get('{%s}fldCharType' % W)
                        if ft == 'separate': ind = True; j += 1; continue
                        elif ft == 'end': eidx = j; break
                    if ind: druns.append(r)
                    j += 1
                if fname and fname in data and druns:
                    for k, dr in enumerate(druns):
                        t = dr.find('{%s}t' % W)
                        if t is not None:
                            t.text = str(data[fname]) if k == 0 else ''
                            t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
                i = eidx + 1; continue
            i += 1
    return b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\r\n" + etree.tostring(tree, encoding='unicode').encode('utf-8')


def detect_fields(docx_path):
    with zipfile.ZipFile(docx_path) as z:
        xml = z.read('word/document.xml')
    tree = etree.fromstring(xml)
    fields = []
    for instr in tree.findall('.//{%s}instrText' % W):
        m = re.search(r'MERGEFIELD\s+(\S+)', instr.text or '')
        if m and m.group(1) not in fields:
            fields.append(m.group(1))
    return fields


def fix_pingshu_xml(xml_bytes):
    tree = etree.fromstring(xml_bytes)
    body = tree.find('{%s}body' % W)
    children = list(body)
    to_remove = []
    for ch in children[1:]:
        tag = ch.tag.split('}')[1]
        if tag == 'p':
            text = ''.join(t.text or '' for t in ch.findall('.//{%s}t' % W)).strip()
            if not text and len(to_remove) < 6:
                to_remove.append(ch)
            else:
                break
        else:
            break
    for ch in to_remove:
        body.remove(ch)

    # 更新 body 末端的 sectPr 頁面設定，不插入帶 sectPr 的段落（避免合併後出現空白頁）
    sect_pr = body.find('{%s}sectPr' % W)
    if sect_pr is None:
        sect_pr = etree.SubElement(body, '{%s}sectPr' % W)

    for tag in ['pgSz', 'pgMar']:
        old = sect_pr.find('{%s}%s' % (W, tag))
        if old is not None:
            sect_pr.remove(old)

    sect_pr.append(etree.fromstring(
        '<w:pgSz xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'
        ' w:w="11906" w:h="16838"/>'
    ))
    sect_pr.append(etree.fromstring(
        '<w:pgMar xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'
        ' w:top="1361" w:right="1134" w:bottom="851" w:left="851"'
        ' w:header="851" w:footer="992" w:gutter="0"/>'
    ))

    for trH in tree.findall('.//{%s}trHeight' % W):
        trH.set('{%s}hRule' % W, 'atLeast')
    tbl = tree.findall('.//{%s}tbl' % W)
    if tbl:
        paras = tbl[0].findall('.//{%s}p' % W)
        for i, p in enumerate(paras):
            pPr = p.find('{%s}pPr' % W)
            if pPr is None: continue
            sp = pPr.find('{%s}spacing' % W)
            if sp is None: continue
            if sp.get('{%s}line' % W) == '480':
                sp.set('{%s}line' % W, '400')
            if i == len(paras) - 2:
                sp.set('{%s}before' % W, '200')
                sp.attrib.pop('{%s}beforeLines' % W, None)
            if i == len(paras) - 1:
                sp.set('{%s}before' % W, '500')
                sp.attrib.pop('{%s}beforeLines' % W, None)
    return b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\r\n" + etree.tostring(tree, encoding='unicode').encode('utf-8')


def shrink_pingyue_xml(xml_bytes):
    tree = etree.fromstring(xml_bytes)
    body = tree.find('{%s}body' % W)
    paras = list(body)[:-1]
    for idx, p in enumerate(paras):
        pPr = p.find('{%s}pPr' % W)
        if pPr is None:
            pPr = etree.SubElement(p, '{%s}pPr' % W)
        sp = pPr.find('{%s}spacing' % W)
        if idx == 0:
            if sp is not None:
                sp.set('{%s}line' % W, '280'); sp.set('{%s}lineRule' % W, 'exact')
            else:
                sp = etree.SubElement(pPr, '{%s}spacing' % W)
                sp.set('{%s}line' % W, '280'); sp.set('{%s}lineRule' % W, 'exact')
            for sz in p.findall('.//{%s}sz' % W): sz.set('{%s}val' % W, '28')
            for szCs in p.findall('.//{%s}szCs' % W): szCs.set('{%s}val' % W, '28')
            for t in p.findall('.//{%s}t' % W):
                if t.text and t.text.strip(): t.text = t.text.strip()
        elif idx in [1, 2]:
            jc = pPr.find('{%s}jc' % W)
            if jc is None: jc = etree.SubElement(pPr, '{%s}jc' % W)
            jc.set('{%s}val' % W, 'right')
            if sp is not None:
                sp.set('{%s}line' % W, '164'); sp.set('{%s}lineRule' % W, 'exact')
            else:
                sp = etree.SubElement(pPr, '{%s}spacing' % W)
                sp.set('{%s}line' % W, '164'); sp.set('{%s}lineRule' % W, 'exact')
            for sz in p.findall('.//{%s}sz' % W): sz.set('{%s}val' % W, '14')
            for szCs in p.findall('.//{%s}szCs' % W): szCs.set('{%s}val' % W, '14')
            for t in p.findall('.//{%s}t' % W):
                if t.text: t.text = t.text.lstrip()
        else:
            if sp is not None:
                sp.set('{%s}line' % W, '164'); sp.set('{%s}lineRule' % W, 'exact')
            else:
                sp = etree.SubElement(pPr, '{%s}spacing' % W)
                sp.set('{%s}line' % W, '164'); sp.set('{%s}lineRule' % W, 'exact')
            for sz in p.findall('.//{%s}sz' % W): sz.set('{%s}val' % W, '14')
            for szCs in p.findall('.//{%s}szCs' % W): szCs.set('{%s}val' % W, '14')
    return b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\r\n" + etree.tostring(tree, encoding='unicode').encode('utf-8')


def merge_docx_files(docx1_path, docx2_path, out_path):
    """將兩個 docx 合併成一個（聘書 + 聘約，加分頁）"""
    with zipfile.ZipFile(docx1_path) as z1:
        orig1 = {n: z1.read(n) for n in z1.namelist()}
    with zipfile.ZipFile(docx2_path) as z2:
        orig2 = {n: z2.read(n) for n in z2.namelist()}

    tree1 = etree.fromstring(orig1['word/document.xml'])
    tree2 = etree.fromstring(orig2['word/document.xml'])

    body1 = tree1.find('{%s}body' % W)
    body2 = tree2.find('{%s}body' % W)

    children2 = list(body2)
    sect_pr = body1.find('{%s}sectPr' % W)

    page_break = etree.fromstring(
        '<w:p xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        '<w:r><w:br w:type="page"/></w:r></w:p>'
    )

    body1_children = list(body1)
    insert_pos = len(body1_children) - 1 if sect_pr is not None else len(body1_children)
    body1.insert(insert_pos, page_break)

    offset = 1
    for child in children2:
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag != 'sectPr':
            body1.insert(insert_pos + offset, child)
            offset += 1

    merged_xml = b"<?xml version='1.0' encoding='UTF-8' standalone='yes'?>\r\n" + etree.tostring(tree1, encoding='unicode').encode('utf-8')

    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n, c in orig1.items():
            zout.writestr(n, merged_xml if n == 'word/document.xml' else c)


def docx_to_pdf(docx_path, out_dir):
    subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'pdf',
         '--outdir', out_dir, docx_path],
        capture_output=True, timeout=120
    )
    base = os.path.splitext(os.path.basename(docx_path))[0]
    return os.path.join(out_dir, base + '.pdf')


def normalize_row(headers, row):
    data = {}
    for h, v in zip(headers, row):
        if h:
            data[h] = str(v) if v is not None else ''
            nk = re.sub(r'\s+', '__', h.strip())
            data[nk] = str(v) if v is not None else ''
    return data


@app.route('/')
def index():
    return send_from_directory('.', 'index.html')


@app.route('/detect-fields', methods=['POST'])
def detect_fields_route():
    if 'docx' not in request.files:
        return jsonify({'error': '請上傳 Word 檔案'}), 400
    f = request.files['docx']
    tmp = tempfile.mkdtemp()
    try:
        path = os.path.join(tmp, 'template.docx')
        f.save(path)
        fields = detect_fields(path)
        return jsonify({'fields': fields})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@app.route('/merge', methods=['POST'])
def merge():
    if 'docx' not in request.files or 'excel' not in request.files:
        return jsonify({'error': '請上傳聘書範本和資料檔案'}), 400

    docx_file    = request.files['docx']
    pingyue_file = request.files.get('pingyue')
    excel_file   = request.files['excel']
    shrink_mode  = request.form.get('shrink', 'true') == 'true'

    tmp = tempfile.mkdtemp()
    try:
        docx_path  = os.path.join(tmp, 'pingshu.docx')
        excel_path = os.path.join(tmp, 'data.xlsx')
        docx_file.save(docx_path)
        excel_file.save(excel_path)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Excel 檔案是空的'}), 400
        headers = [str(h).strip() if h else '' for h in rows[0]]
        data_rows = [
            normalize_row(headers, r)
            for r in rows[1:]
            if any(v is not None and str(v).strip() for v in r)
        ]
        if not data_rows:
            return jsonify({'error': '找不到有效資料列'}), 400

        with zipfile.ZipFile(docx_path) as z:
            ps_orig = {n: z.read(n) for n in z.namelist()}

        ps_fixed_base = fix_pingshu_xml(ps_orig['word/document.xml'])

        work_dir = os.path.join(tmp, 'work')
        os.makedirs(work_dir)

        py_reader = None
        if pingyue_file and pingyue_file.filename:
            py_path = os.path.join(tmp, 'pingyue.docx')
            pingyue_file.save(py_path)
            with zipfile.ZipFile(py_path) as z:
                py_orig = {n: z.read(n) for n in z.namelist()}
            py_xml = shrink_pingyue_xml(py_orig['word/document.xml']) if shrink_mode else py_orig['word/document.xml']
            py_docx = os.path.join(work_dir, 'pingyue.docx')
            with zipfile.ZipFile(py_docx, 'w', zipfile.ZIP_DEFLATED) as zout:
                for n, c in py_orig.items():
                    zout.writestr(n, py_xml if n == 'word/document.xml' else c)
            py_pdf = docx_to_pdf(py_docx, work_dir)
            if os.path.exists(py_pdf):
                py_reader = PdfReader(py_pdf)

        fw = PdfWriter()
        for idx, row in enumerate(data_rows):
            merged_xml = replace_mergefields(ps_fixed_base, row)
            dp = os.path.join(work_dir, f'ps_{idx+1}.docx')
            with zipfile.ZipFile(dp, 'w', zipfile.ZIP_DEFLATED) as zout:
                for n, c in ps_orig.items():
                    zout.writestr(n, merged_xml if n == 'word/document.xml' else c)
            pdf_path = docx_to_pdf(dp, work_dir)
            if os.path.exists(pdf_path):
                r2 = PdfReader(pdf_path)
                fw.add_page(r2.pages[0])
                if py_reader:
                    for pg in py_reader.pages:
                        fw.add_page(pg)

        out_path = os.path.join(tmp, '合併列印.pdf')
        with open(out_path, 'wb') as f:
            fw.write(f)

        return send_file(
            out_path,
            mimetype='application/pdf',
            as_attachment=True,
            download_name='合併列印.pdf'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@app.route('/merge-word', methods=['POST'])
def merge_word():
    if 'docx' not in request.files or 'excel' not in request.files:
        return jsonify({'error': '請上傳聘書範本和資料檔案'}), 400

    docx_file    = request.files['docx']
    pingyue_file = request.files.get('pingyue')
    excel_file   = request.files['excel']
    shrink_mode  = request.form.get('shrink', 'true') == 'true'

    tmp = tempfile.mkdtemp()
    try:
        docx_path  = os.path.join(tmp, 'pingshu.docx')
        excel_path = os.path.join(tmp, 'data.xlsx')
        docx_file.save(docx_path)
        excel_file.save(excel_path)

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return jsonify({'error': 'Excel 檔案是空的'}), 400
        headers = [str(h).strip() if h else '' for h in rows[0]]
        data_rows = [
            normalize_row(headers, r)
            for r in rows[1:]
            if any(v is not None and str(v).strip() for v in r)
        ]
        if not data_rows:
            return jsonify({'error': '找不到有效資料列'}), 400

        with zipfile.ZipFile(docx_path) as z:
            ps_orig = {n: z.read(n) for n in z.namelist()}

        ps_fixed_base = fix_pingshu_xml(ps_orig['word/document.xml'])

        # 處理聘約
        py_orig = None
        py_xml_bytes = None
        if pingyue_file and pingyue_file.filename:
            py_path = os.path.join(tmp, 'pingyue.docx')
            pingyue_file.save(py_path)
            with zipfile.ZipFile(py_path) as z:
                py_orig = {n: z.read(n) for n in z.namelist()}
            py_xml_bytes = shrink_pingyue_xml(py_orig['word/document.xml']) if shrink_mode else py_orig['word/document.xml']

        zip_path = os.path.join(tmp, '教師聘書合併.zip')
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zout_all:
            for idx, row in enumerate(data_rows):
                name = row.get('姓名') or row.get('教師名稱') or f'第{idx+1}位'
                safe_name = re.sub(r'[\\/:*?"<>|]', '', name)
                merged_xml = replace_mergefields(ps_fixed_base, row)

                ps_docx = os.path.join(tmp, f'ps_{idx+1}.docx')
                with zipfile.ZipFile(ps_docx, 'w', zipfile.ZIP_DEFLATED) as zdoc:
                    for n, c in ps_orig.items():
                        zdoc.writestr(n, merged_xml if n == 'word/document.xml' else c)

                if py_orig and py_xml_bytes:
                    py_docx = os.path.join(tmp, f'py_{idx+1}.docx')
                    with zipfile.ZipFile(py_docx, 'w', zipfile.ZIP_DEFLATED) as zdoc:
                        for n, c in py_orig.items():
                            zdoc.writestr(n, py_xml_bytes if n == 'word/document.xml' else c)
                    merged_docx = os.path.join(tmp, f'{idx+1}_{safe_name}.docx')
                    merge_docx_files(ps_docx, py_docx, merged_docx)
                    zout_all.write(merged_docx, f'{idx+1}_{safe_name}.docx')
                else:
                    zout_all.write(ps_docx, f'{idx+1}_{safe_name}.docx')

        return send_file(
            zip_path,
            mimetype='application/zip',
            as_attachment=True,
            download_name='教師聘書合併.zip'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
